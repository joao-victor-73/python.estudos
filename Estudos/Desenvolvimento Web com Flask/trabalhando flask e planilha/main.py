# Importando as bibliotecas

from flask import Flask, render_template

import openpyxl
# /\ Essa biblioteca servirá para trabalhar com o arquivo xlsx(a planilha)

app = Flask(__name__)


"""
                                                OBS

>>> Pelo que venho estudando, o melhor procedimento aqui é converter os dados da planilha para 
    arquivo JSON e então passar para o html, podendo de lá, fazer todo o tratamento e visulização
    dos dados fornecidos, e também podendo fazer a inserção dos dados em uma tabela SQL (de preferencia
    o MySQL)."
"""


# Variável para o caminho do arquivo
CAMINHO_ARQUIVO = "D:\\MeusRepositorios\\Pyhton\\Estudos\\Desenvolvimento Web com Flask\\trabalhando flask e planilha\\ListaCompleta.xlsx"


# Função para ler os dados da planilha
def ler_dados_da_planilha():
    workbook = openpyxl.load_workbook(CAMINHO_ARQUIVO)
    sheet = workbook.active
    dados = []

    # cabecalho = []

    # Obter o cabeçalho de cada coluna
    cabecalho = [cell.value for cell in next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))]

    # Itera sobre as linhas da planilha, pulando a linha do cabeçalho;
    for row in sheet.iter_rows(min_row=4, values_only=True):
        
        # Cria um dicionário com os dados da linha
        row_dicionario = dict(zip(cabecalho, row))
        """
         a função zip() vai ser útil para combinar os cabeçalhos das colunas (cabecalho) 
         com os dados da linha atual (row). Isso cria pares chave-valor.
        """

        dados.append(row_dicionario)


        # print(f"Linha: {row}")

        """        
        if row[0] == 'PLANILHA DE JOGOS':
            cabecalho.append(row[0])
            # print(f"{row[0]} adicionado com sucesso!")

        elif row[0] == 'Nº':
            cabecalho.append(row[0:7])
            # print(f"{row[0:7]} adicionado com sucesso!")

        # Verifica se há células vazias na linha
        if None != row:
            dados.append(row)
        """

    # print(f"LISTA CABECALHO: {cabecalho}")

    """
    for row in sheet.iter_rows(values_only=True):
        cont += 1
        if row != 'None':
            dados.append(row)
    """

    return dados, cabecalho


# Rota para a página inicial
@app.route('/')
def principal():
    dados, cabecalho = ler_dados_da_planilha()

    return render_template('index.html', dados=dados, cabecalho=cabecalho)


if __name__ == '__main__':
    app.run(debug=True)
