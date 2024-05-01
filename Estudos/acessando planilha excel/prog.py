import openpyxl

# Abre o arquivo Excel
workbook = openpyxl.load_workbook("gam'ers.xlsx")

# Seleciona a planilha ativa (por padrão, a primeira planilha é selecionada)
sheet_ativa = workbook.active

# Seleciona uma planilha específica pelo nome
nome_da_planilha = 'Lista_Completa'
sheet_especifica = workbook[nome_da_planilha]

# Vamos tentar acessar uma linha especifica do arquivo:

# Número da linha que você deseja acessar
numero_da_linha = 7  # Por exemplo, acessar a linha 3

# Itera sobre as linhas até chegar na linha desejada
for row in sheet_especifica.iter_rows(min_row=numero_da_linha, max_row=numero_da_linha, values_only=True):
    print(row)

"""
# Itera sobre as linhas e colunas da planilha ativa
for row in sheet_ativa.iter_rows(values_only=True):
    print(row)
"""